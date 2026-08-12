import io
import openpyxl
from typing import Optional

def try_parse_bank_excel(file_bytes: bytes, file_type: str) -> Optional[int]:
    """
    Attempts to parse a bank account statement Excel file directly.
    For bank_income: Sums the amount where '구분' column equals '무통장입금'.
    For bank_expense: Sums the amount where '내 통장 표시' column contains '취소환불'.
    Returns the sum as int, or None if the sheet format doesn't match.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        
        header_row = -1
        col_mapping = {}
        
        for r in range(1, min(31, sheet.max_row + 1)):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            row_str_vals = [str(v).strip() if v is not None else "" for v in row_vals]
            
            if "거래일시" in row_str_vals or "구분" in row_str_vals:
                header_row = r
                for idx, val in enumerate(row_str_vals):
                    if val:
                        col_mapping[val.strip()] = idx + 1
                break
                
        if header_row == -1:
            return None
            
        gubun_col = col_mapping.get("구분")
        memo_col = col_mapping.get("메모")
        my_bank_col = col_mapping.get("내 통장 표시") or col_mapping.get("기재내용")
        
        amt_col = None
        if file_type == 'bank_income':
            for name, col_idx in col_mapping.items():
                if '입금' in name:
                    amt_col = col_idx
                    break
        elif file_type == 'bank_expense':
            for name, col_idx in col_mapping.items():
                if '출금' in name or '지급' in name:
                    amt_col = col_idx
                    break
                    
        if amt_col is None:
            return None
            
        total_sum = 0
        for r in range(header_row + 1, sheet.max_row + 1):
            match = False
            if file_type == 'bank_income':
                if gubun_col is not None:
                    gubun_val = sheet.cell(row=r, column=gubun_col).value
                    if gubun_val is not None and str(gubun_val).strip() == "무통장입금":
                        match = True
                if not match and memo_col is not None:
                    memo_val = sheet.cell(row=r, column=memo_col).value
                    if memo_val is not None and str(memo_val).strip() == "무통장입금":
                        match = True
            elif file_type == 'bank_expense':
                if my_bank_col is not None:
                    my_bank_val = sheet.cell(row=r, column=my_bank_col).value
                    if my_bank_val is not None and "취소환불" in str(my_bank_val).strip():
                        match = True
                        
            if match:
                amt_val = sheet.cell(row=r, column=amt_col).value
                if amt_val is not None:
                    try:
                        if isinstance(amt_val, (int, float)):
                            val = int(amt_val)
                        else:
                            cleaned_val = str(amt_val).replace(",", "").replace("₩", "").strip()
                            val = int(float(cleaned_val))
                        total_sum += val
                    except Exception:
                        pass
                        
        return total_sum
    except Exception as e:
        print(f"Error in try_parse_bank_excel: {e}")
        return None

def try_parse_toss_excel(file_bytes: bytes, classification: str) -> Optional[int]:
    """
    Attempts to parse Toss Payments Excel file.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb['매출'] if '매출' in wb.sheetnames else wb.active
        
        search_key = None
        if classification in ('신용/체크카드', '신용·체크카드'):
            search_key = '신용·체크카드 합계'
        elif classification == '현금영수증 자동발행':
            search_key = '현금영수증 자동발행 합계'
        elif classification == '현금영수증 별도발급':
            search_key = '현금영수증 별도발급 합계'
        elif classification in ('기타 합계', '기타(정규영수증 외 매출분)'):
            search_key = '기타 합계'
            
        if not search_key:
            return None
            
        header_row = -1
        amt_col = None
        
        for r in range(1, 10):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            row_str_vals = [str(v).strip() if v is not None else "" for v in row_vals]
            if "합계 (결제액-취소완료액)" in row_str_vals:
                header_row = r
                amt_col = row_str_vals.index("합계 (결제액-취소완료액)") + 1
                break
                
        if header_row == -1 or amt_col is None:
            amt_col = 7
            header_row = 1
            
        def clean_str(s):
            return str(s).replace(" ", "").replace("·", "").replace("•", "").strip()
            
        clean_search_key = clean_str(search_key)
        
        for r in range(header_row + 1, sheet.max_row + 1):
            col2_val = sheet.cell(row=r, column=2).value
            if col2_val is not None:
                clean_col2 = clean_str(col2_val)
                if clean_search_key in clean_col2:
                    amt_val = sheet.cell(row=r, column=amt_col).value
                    if amt_val is not None:
                        try:
                            if isinstance(amt_val, (int, float)):
                                return int(amt_val)
                            else:
                                cleaned_val = str(amt_val).replace(",", "").replace("₩", "").strip()
                                return int(float(cleaned_val))
                        except Exception:
                            pass
        return None
    except Exception as e:
        print(f"Error in try_parse_toss_excel: {e}")
        return None

def try_parse_naver_pay_excel(file_bytes: bytes, classification: str) -> Optional[int]:
    """
    Attempts to parse Naver Pay / Smartstore Excel file.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb['부가세신고 월별내역'] if '부가세신고 월별내역' in wb.sheetnames else wb.active
        
        header_row = -1
        col_mapping = {}
        for r in range(1, 6):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            row_str_vals = [str(v).strip() if v is not None else "" for v in row_vals]
            if "부가세 신고기간" in row_str_vals or "신용카드 매출전표" in row_str_vals:
                header_row = r
                for idx, val in enumerate(row_str_vals):
                    if val:
                        col_mapping[val.strip()] = idx + 1
                break
                
        if header_row == -1:
            return None
            
        target_col = col_mapping.get(classification)
        if target_col is None:
            return None
            
        total_sum = 0
        for r in range(header_row + 1, sheet.max_row + 1):
            period_val = sheet.cell(row=r, column=1).value
            if period_val is None or str(period_val).strip() == "":
                break
            
            val = sheet.cell(row=r, column=target_col).value
            if val is not None:
                try:
                    if isinstance(val, (int, float)):
                        total_sum += int(val)
                    else:
                        cleaned_val = str(val).replace(",", "").replace("₩", "").strip()
                        total_sum += int(float(cleaned_val))
                except Exception:
                    pass
        return total_sum
    except Exception as e:
        print(f"Error in try_parse_naver_pay_excel: {e}")
        return None

def try_parse_koces_excel(file_bytes: bytes, classification: str) -> Optional[int]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        
        a1_val = sheet.cell(row=1, column=1).value
        if not a1_val or "KOCES" not in str(a1_val):
            return None
            
        target_row = -1
        for r in range(12, 22):
            g_val = sheet.cell(row=r, column=7).value
            if g_val and "매출총액" in str(g_val):
                target_row = r
                break
                
        if target_row == -1:
            return None
            
        if classification == "카드결제":
            card_val = sheet.cell(row=target_row, column=16).value or 0
            return int(card_val)
            
        elif classification == "간편결제":
            easy_val = sheet.cell(row=target_row, column=25).value or 0
            return int(easy_val)
            
        elif classification == "현금영수증":
            receipt_val = sheet.cell(row=target_row, column=34).value or 0
            return int(receipt_val)
            
        return None
    except Exception as e:
        print(f"Error parsing KOCES excel: {e}")
        return None

def try_parse_qoo10_excel(file_bytes: bytes) -> Optional[int]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        
        a1_val = sheet.cell(row=1, column=1).value
        if not a1_val or "우리은행" not in str(a1_val):
            return None
            
        header_row = -1
        col_mapping = {}
        for r in range(1, min(31, sheet.max_row + 1)):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
            row_str_vals = [str(v).strip() if v is not None else "" for v in row_vals]
            if "거래일시" in row_str_vals or "구분" in row_str_vals:
                header_row = r
                for idx, val in enumerate(row_str_vals):
                    if val:
                        col_mapping[val.strip()] = idx + 1
                break
                
        if header_row == -1:
            return None
            
        content_col = col_mapping.get("기재내용")
        
        amt_col = None
        for name, col_idx in col_mapping.items():
            if '입금' in name:
                amt_col = col_idx
                break
                
        if content_col is None or amt_col is None:
            return None
            
        total_sum = 0
        for r in range(header_row + 1, sheet.max_row + 1):
            content_val = sheet.cell(row=r, column=content_col).value
            if content_val is not None and "EBAY" in str(content_val).upper():
                deposit_val = sheet.cell(row=r, column=amt_col).value
                if deposit_val is not None:
                    try:
                        if isinstance(deposit_val, (int, float)):
                            val = int(deposit_val)
                        else:
                            cleaned_val = str(deposit_val).replace(",", "").replace("₩", "").strip()
                            val = int(float(cleaned_val))
                        total_sum += val
                    except Exception:
                        pass
        return total_sum
    except Exception as e:
        print(f"Error parsing Qoo10 excel: {e}")
        return None

def try_parse_paypal_excel(file_bytes: bytes) -> Optional[tuple[int, str]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        
        headers = [sheet.cell(row=1, column=c).value for c in range(1, 15)]
        headers_str = [str(h).strip() if h is not None else "" for h in headers]
        if "순액" not in headers_str:
            return None
            
        net_col_idx = headers_str.index("순액") + 1
        
        total_usd = 0.0
        for r in range(2, sheet.max_row + 1):
            val = sheet.cell(row=r, column=net_col_idx).value
            if val is not None:
                try:
                    total_usd += float(val)
                except Exception:
                    pass
                    
        exchange_rate = 1480.0
        total_krw = int(round(total_usd * exchange_rate))
        
        reason = f"PayPal 엑셀 분석 결과: 총 순액 ${total_usd:,.2f} USD * 환율 {exchange_rate:,.1f}원 = {total_krw:,}원 (기준 환율: 1,480원 적용)"
        
        return total_krw, reason
    except Exception as e:
        print(f"Error parsing PayPal excel: {e}")
        return None

def try_parse_youtube_excel(file_bytes: bytes) -> Optional[tuple[int, str]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        
        headers = [sheet.cell(row=1, column=c).value for c in range(1, 5)]
        headers_str = [str(h).strip() if h is not None else "" for h in headers]
        if "설명" not in headers_str:
            return None
            
        desc_col_idx = headers_str.index("설명") + 1
        
        amt_col_idx = None
        for idx, h in enumerate(headers_str):
            if "금액" in h:
                amt_col_idx = idx + 1
                break
                
        if amt_col_idx is None:
            return None
            
        total_usd = 0.0
        for r in range(2, sheet.max_row + 1):
            desc_val = sheet.cell(row=r, column=desc_col_idx).value
            amt_val = sheet.cell(row=r, column=amt_col_idx).value
            
            if desc_val is not None and str(desc_val).strip() == "수입 - YouTube":
                if amt_val is not None:
                    try:
                        amt_str = str(amt_val).replace("−", "-").replace(",", "").strip()
                        total_usd += float(amt_str)
                    except Exception:
                        pass
                        
        exchange_rate = 1480.0
        total_krw = int(round(total_usd * exchange_rate))
        
        reason = f"YouTube AdSense 엑셀 분석 결과: 총 수입 ${total_usd:,.2f} USD * 환율 {exchange_rate:,.1f}원 = {total_krw:,}원 (기준 환율: 1,480원 적용)"
        
        return total_krw, reason
    except Exception as e:
        print(f"Error parsing YouTube excel: {e}")
        return None
