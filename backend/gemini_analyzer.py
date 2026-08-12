import os
import re
import io
import openpyxl
from google import genai
from google.genai import types

def excel_to_markdown_openpyxl(file_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    text = ""
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text += f"\n--- 시트명: {sheet_name} ---\n"
        
        max_r = sheet.max_row
        max_c = sheet.max_column
        
        rows = []
        for r in range(1, min(max_r + 1, 300)):  # limit to 300 rows to prevent token size issues
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, max_c + 1)]
            if any(val is not None and str(val).strip() != "" for val in row_vals):
                cleaned_vals = [str(val).replace('\n', ' ').strip() if val is not None else "" for val in row_vals]
                rows.append(cleaned_vals)
                
        if not rows:
            text += "(빈 시트)\n"
            continue
            
        headers = rows[0]
        text += "| " + " | ".join(headers) + " |\n"
        text += "| " + " | ".join(["---"] * len(headers)) + " |\n"
        for row in rows[1:]:
            text += "| " + " | ".join(row) + " |\n"
    return text

def analyze_document_with_gemini(
    api_key: str, 
    file_bytes: bytes, 
    file_name: str, 
    mime_type: str, 
    brand: str, 
    pg_store: str, 
    classification: str, 
    reference: str
) -> dict:
    if not api_key:
        return {"status": "error", "error": "Gemini API 키가 설정되지 않았습니다.", "amount": 0}

    try:
        client = genai.Client(api_key=api_key)
    except Exception as e:
        return {"status": "error", "error": f"Gemini Client 초기화 실패: {str(e)}", "amount": 0}
    
    contents = []
    
    prompt = f"""
역할: 대한민국 전문 세무/회계 매출 검토 AI 비서.
목표: 제공된 증빙 자료(이미지 캡처 또는 엑셀 시트 데이터)를 정밀 분석하여, 아래 지정된 [분석 대상 항목]의 부가세 신고용 최종 금액(원화)을 정확하게 추출하십시오.

[분석 대상 항목 정보]
- 브랜드: {brand or "N/A"}
- PG / 스토어: {pg_store or "N/A"}
- 구분: {classification or "N/A"}
- 참고 기준 (금액 산정 조건): {reference or "N/A"}

[작업 지침 및 규칙]
1. 참고 기준({reference})에 명시된 특정 거래 유형(예: 신용카드 매출, 현금영수증 발행, 기타 등)에 해당하는 숫자를 증빙 자료에서 찾아내십시오.
2. 엑셀 또는 이미지 내의 여러 테이블/합계 중에서 반드시 해당 브랜드와 PG/스토어에 맞는 항목을 분별하십시오.
3. 결과는 설명 없이 **오직 부호를 포함한 정수 숫자**로만 답변해 주십시오. (예: 330579843 또는 -110436)
4. 금액에 포함된 천 단위 콤마(,), 원(₩) 기호, 한글 단어 등은 모두 제거하고 순수 정수 값만 반환해야 합니다.
5. 취소나 환불 등으로 마이너스 금액이 명시되어 있거나, 참고 조건에 부합하는 계산 결과가 음수라면 반드시 마이너스 부호(-)를 붙이십시오.
6. 만약 자료에 해당 금액이 존재하지 않거나, 대상을 식별할 수 없는 경우 0을 반환하십시오.
7. 답변에는 어떤 설명, 인사말, 단위 설명도 작성하지 마십시오. 오직 숫자만 작성하십시오.
8. 만약 자료(예: 이미지)에 월별 내역과 '총 합계' 혹은 '합계'가 함께 존재한다면, 개별 월별 금액이 아닌 '총 합계' 혹은 '합계' 행의 금액을 최종 금액으로 추출해야 합니다.
"""
    contents.append(prompt)
    
    if mime_type.startswith("image/") or file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif')):
        actual_mime = mime_type
        if "octet-stream" in mime_type or not mime_type:
            if file_name.lower().endswith('.png'): actual_mime = 'image/png'
            elif file_name.lower().endswith(('.jpg', '.jpeg')): actual_mime = 'image/jpeg'
            elif file_name.lower().endswith('.webp'): actual_mime = 'image/webp'
            else: actual_mime = 'image/png'
            
        try:
            image_part = types.Part.from_bytes(data=file_bytes, mime_type=actual_mime)
            contents.append(image_part)
        except Exception as e:
            return {"status": "error", "error": f"이미지 데이터 변환 실패: {str(e)}", "amount": 0}
    else:
        try:
            excel_text = excel_to_markdown_openpyxl(file_bytes)
            contents.append(f"\n[업로드된 엑셀 데이터]\n{excel_text}")
        except Exception as e:
            return {"status": "error", "error": f"엑셀 파일 텍스트 변환 실패: {str(e)}", "amount": 0}
            
    try:
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=contents
        )
        response_text = response.text.strip()
        
        cleaned = re.sub(r'[, \s가-힣원₩]', '', response_text)
        match = re.search(r'-?\d+', cleaned)
        if match:
            amount = int(match.group(0))
            return {
                "status": "success",
                "amount": amount,
                "raw_response": response_text
            }
        else:
            return {
                "status": "error",
                "error": f"텍스트 내에서 금액 숫자를 식별하지 못했습니다. (추출된 텍스트: {response_text})",
                "amount": 0,
                "raw_response": response_text
            }
    except Exception as e:
        return {"status": "error", "error": f"Gemini API 호출 실패: {str(e)}", "amount": 0}
