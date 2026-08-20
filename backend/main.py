import os
import io
import re
import uuid
import shutil
import openpyxl
from typing import Optional, List, Dict, Any

from dotenv import load_dotenv

# Load parent Dashboard/.env and local backend/.env
parent_env = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".env"))
if os.path.exists(parent_env):
    load_dotenv(parent_env)
load_dotenv()

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from supabase import create_client, Client
from pydantic import BaseModel

# Import custom modular helpers
from gemini_analyzer import analyze_document_with_gemini
from excel_parsers import (
    try_parse_bank_excel,
    try_parse_toss_excel,
    try_parse_naver_pay_excel,
    try_parse_koces_excel,
    try_parse_qoo10_excel,
    try_parse_paypal_excel,
    try_parse_youtube_excel,
)
from report_exporter import export_report_excel
from production_calculator import parse_production_excel, parse_shipment_notes_excel


app = FastAPI(title="VAT Report Automation API")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:3005",
        "http://127.0.0.1:3005",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Supabase Credentials
SUPABASE_URL = os.getenv("SUPABASE_URL") or os.getenv("VITE_SUPABASE_URL", "https://uxvifhjwdsalugzvogds.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY") or os.getenv("VITE_SUPABASE_ANON_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InV4dmlmaGp3ZHNhbHVnenZvZ2RzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODQyNDkzMDksImV4cCI6MjA5OTgyNTMwOX0.-BHD4cuq6hDNUoqsmLpx-VrubtMUdglrA0Upujz4J-M")

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Storage directories
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "sample", "부가세신고 매출자료 정리.xlsx")


class CreateReportPayload(BaseModel):
    year: int
    quarter: int


class RowUpdatePayload(BaseModel):
    amount: Optional[int] = None
    memo: Optional[str] = None
    status: Optional[str] = None


class SettingPayload(BaseModel):
    value: str


def get_gemini_key() -> Optional[str]:
    env_key = os.getenv("GEMINI_API_KEY")
    if env_key:
        return env_key
    try:
        res = supabase.table("settings").select("value").eq("key", "gemini_api_key").execute()
        if res.data:
            return res.data[0]["value"]
    except Exception as e:
        print(f"Error fetching API Key: {e}")
    return None


def recalculate_bank_transfer(report_id: str, brand: str):
    try:
        res = supabase.table("report_rows").select("*").eq("report_id", report_id).eq("brand", brand).execute()
        rows = res.data or []
        
        bank_row = None
        for r in rows:
            if r["classification"] == "무통장입금":
                bank_row = r
                break
                
        if not bank_row:
            return

        toss_receipt_row = None
        for r in rows:
            if r["pg_store"] == "토스페이먼츠" and r["classification"] == "현금영수증 별도발급":
                toss_receipt_row = r
                break
                
        bank_toss_val = toss_receipt_row["amount"] if (toss_receipt_row and toss_receipt_row["amount"] is not None) else 0

        bank_inc = bank_row.get("bank_income") or 0
        bank_exp = bank_row.get("bank_expense") or 0
        bank_koces = bank_row.get("bank_koces_receipt") or 0

        calculated_amt = bank_inc - bank_exp - bank_toss_val - (bank_koces if brand == "오터" else 0)

        supabase.table("report_rows").update({
            "amount": calculated_amt,
            "bank_toss_receipt": bank_toss_val
        }).eq("id", bank_row["id"]).execute()

    except Exception as e:
        print(f"Error recalculating bank transfer row for {brand}: {e}")


def initialize_report_rows(report_id: str):
    if not os.path.exists(TEMPLATE_PATH):
        print(f"Template path does not exist: {TEMPLATE_PATH}")
        return

    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
        sheet = wb.active

        rows_to_insert = []
        current_brand = None
        current_pg = None

        for r in range(3, 81):
            col_a = sheet.cell(row=r, column=1).value
            col_b = sheet.cell(row=r, column=2).value
            col_c = sheet.cell(row=r, column=3).value
            col_d = sheet.cell(row=r, column=4).value
            col_e = sheet.cell(row=r, column=5).value

            if col_a and str(col_a).strip():
                current_brand = str(col_a).strip()
            if col_b and str(col_b).strip():
                current_pg = str(col_b).strip()

            classification = str(col_c).strip() if col_c else ""

            if not current_brand and not classification:
                continue

            ref_parts = []
            if col_d is not None:
                val_d_str = str(col_d).strip()
                if val_d_str.startswith('='):
                    ref_parts.append(f"수식: {val_d_str}")
            if col_e is not None and str(col_e).strip():
                ref_parts.append(f"참고: {str(col_e).strip()}")

            reference = " | ".join(ref_parts)

            status = "empty"
            if classification == "합계" or (current_brand == "판매금액 총합"):
                status = "formula"

            rows_to_insert.append({
                "report_id": report_id,
                "brand": current_brand or "",
                "pg_store": current_pg or "",
                "classification": classification,
                "amount": None,
                "reference": reference,
                "excel_row": r,
                "status": status,
                "memo": ""
            })

        if rows_to_insert:
            supabase.table("report_rows").insert(rows_to_insert).execute()
    except Exception as e:
        print(f"Error initializing report rows: {e}")


@app.get("/api/reports")
def get_all_reports():
    try:
        res = supabase.table("reports").select("*").order("year", desc=True).order("quarter", desc=True).execute()
        return res.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/reports")
def create_report(payload: CreateReportPayload):
    try:
        check_res = supabase.table("reports").select("*").eq("year", payload.year).eq("quarter", payload.quarter).execute()
        if check_res.data:
            raise HTTPException(status_code=400, detail=f"{payload.year}년 {payload.quarter}분기 보고서가 이미 존재합니다.")

        insert_res = supabase.table("reports").insert({
            "year": payload.year,
            "quarter": payload.quarter,
            "status": "in_progress"
        }).execute()
        report = insert_res.data[0]

        initialize_report_rows(report["id"])
        return report
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/reports/{report_id}")
def get_report_details(report_id: str):
    try:
        report_res = supabase.table("reports").select("*").eq("id", report_id).execute()
        if not report_res.data:
            raise HTTPException(status_code=404, detail="Report not found")
        report = report_res.data[0]

        rows_res = supabase.table("report_rows").select("*").eq("report_id", report_id).order("excel_row", desc=False).execute()
        rows = rows_res.data

        return {
            "report": report,
            "rows": rows
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/rows/{row_id}")
def update_report_row(row_id: str, payload: RowUpdatePayload):
    try:
        row_res = supabase.table("report_rows").select("*").eq("id", row_id).execute()
        if not row_res.data:
            raise HTTPException(status_code=404, detail="Row not found")
        row = row_res.data[0]

        update_data = {}
        if payload.amount is not None:
            update_data["amount"] = payload.amount
            update_data["status"] = "success"
        if payload.memo is not None:
            update_data["memo"] = payload.memo
        if payload.status is not None:
            update_data["status"] = payload.status

        if not update_data:
            return row

        supabase.table("report_rows").update(update_data).eq("id", row_id).execute()

        if row["classification"] == "무통장입금":
            recalculate_bank_transfer(row["report_id"], row["brand"])
        elif row["pg_store"] == "토스페이먼츠" and row["classification"] == "현금영수증 별도발급":
            recalculate_bank_transfer(row["report_id"], row["brand"])

        updated_row_res = supabase.table("report_rows").select("*").eq("id", row_id).execute()
        return updated_row_res.data[0]
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==============================================================================
# Production Order & Batch Calculator Endpoints
# ==============================================================================

class ProductCatalogPayload(BaseModel):
    name: str
    category: str = "삼각"
    batch_size: int = 8
    min_bumper_qty: int = 0
    is_confirmed: bool = True
    parent_scone_name: Optional[str] = None
    oven_number: Optional[str] = "1"
    heavy_cream_per_panel: Optional[int] = 0
    sort_order: Optional[int] = 999


class ReorderItemPayload(BaseModel):
    name: str
    sort_order: int
    is_separator: Optional[bool] = False

class ReorderCatalogPayload(BaseModel):
    orders: List[ReorderItemPayload]


class SetItemComponentPayload(BaseModel):
    product_name: str
    quantity: int = 1

class SetCatalogPayload(BaseModel):
    id: Optional[str] = None
    name: Optional[str] = None
    set_name: Optional[str] = None
    description: Optional[str] = ""
    is_confirmed: bool = True
    items: Optional[List[SetItemComponentPayload]] = None
    components: Optional[List[SetItemComponentPayload]] = None

class ProductionRecordSavePayload(BaseModel):
    record_date: Optional[str] = None
    records: list


@app.post("/api/production/parse-excel")
async def api_parse_production_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        
        # 1. Build set_catalog_map from DB
        set_catalog_map = {}
        try:
            sets_res = supabase.table("set_catalog").select("*").execute()
            if sets_res.data:
                set_ids = [s["id"] for s in sets_res.data]
                items_res = supabase.table("set_items").select("*").in_("set_id", set_ids).execute()
                items_by_set = {}
                if items_res.data:
                    for it in items_res.data:
                        items_by_set.setdefault(it["set_id"], []).append({
                            "product_name": it["product_name"],
                            "quantity": it.get("quantity", 1)
                        })
                
                for s in sets_res.data:
                    set_catalog_map[s["name"]] = {
                        "id": s["id"],
                        "name": s["name"],
                        "items": items_by_set.get(s["id"], [])
                    }
        except Exception as e:
            print("Failed to load set_catalog for excel parsing:", e)

        # 2. Check single product DB catalog overrides
        catalog_res = supabase.table("product_catalog").select("*").execute()
        catalog_map = {}
        if catalog_res.data:
            for c in catalog_res.data:
                catalog_map[c["name"]] = c

        # 3. Parse excel with set decomposition & catalog map
        parse_result = parse_production_excel(contents, set_catalog_map=set_catalog_map, catalog_dict=catalog_map)
        parsed_items = parse_result.get("items", [])
        set_breakdowns = parse_result.get("set_breakdowns", [])

        # Auto-register new unconfirmed products & override settings
        new_products_to_insert = []
        for item in parsed_items:
            name = item["product_name"]
            if name in catalog_map:
                item["category"] = catalog_map[name]["category"]
                item["batch_size"] = catalog_map[name]["batch_size"]
                item["min_bumper_qty"] = catalog_map[name].get("min_bumper_qty", 0)
                item["is_confirmed"] = catalog_map[name].get("is_confirmed", True)
                item["parent_scone_name"] = catalog_map[name].get("parent_scone_name")
                item["oven_number"] = catalog_map[name].get("oven_number", "1")
                item["heavy_cream_per_panel"] = catalog_map[name].get("heavy_cream_per_panel", 0)
                item["sort_order"] = catalog_map[name].get("sort_order", 999)
            else:
                item["min_bumper_qty"] = item.get("min_bumper_qty", 0)
                item["is_confirmed"] = False
                item["parent_scone_name"] = None
                item["oven_number"] = "1"
                item["heavy_cream_per_panel"] = 0
                item["sort_order"] = 999
                
                # Do not insert into product_catalog if it is already a registered Set Product!
                if name not in set_catalog_map:
                    new_products_to_insert.append({
                        "name": name,
                        "category": item["category"],
                        "batch_size": item["batch_size"],
                        "min_bumper_qty": 0,
                        "is_confirmed": False,
                        "parent_scone_name": None,
                        "oven_number": "1",
                        "heavy_cream_per_panel": 0,
                        "sort_order": 999,
                    })

        # Insert newly detected products as unconfirmed into DB
        if new_products_to_insert:
            try:
                supabase.table("product_catalog").upsert(new_products_to_insert, on_conflict="name").execute()
            except Exception as e:
                print("Failed to auto-insert new catalog items:", e)

        # Recalculate required_qty & production_qty & bumper logic cleanly for each category
        for item in parsed_items:
            import math
            order_qty = item.get("order_qty", 0)
            extra_qty = item.get("extra_qty", 0)
            carryover_qty = item.get("carryover_qty", 0)
            
            required_qty = order_qty + extra_qty
            prod_qty = max(0, required_qty - carryover_qty)
            bs = item.get("batch_size", 8)
            extra_panels = item.get("min_bumper_qty", 0) or 0
            cat = item.get("category")

            item["required_qty"] = required_qty
            item["production_qty"] = prod_qty

            if cat == '스틱':
                base_panels = math.ceil(prod_qty / 9.0) if prod_qty > 0 else 0
                item["panels"] = base_panels
                item["is_bumper_applied"] = False
                item["base_panels"] = base_panels
                item["excess_qty"] = (base_panels * 9) - prod_qty
            elif cat == '미니쉐이크':
                base_panels = math.ceil(prod_qty / 4.0) if prod_qty > 0 else 0
                item["panels"] = base_panels
                item["is_bumper_applied"] = False
                item["base_panels"] = base_panels
                item["excess_qty"] = (base_panels * 4) - prod_qty
            elif cat == '미니큐브':
                base_panels = math.ceil(prod_qty / 2.0) if prod_qty > 0 else 0
                item["panels"] = base_panels
                item["is_bumper_applied"] = False
                item["base_panels"] = base_panels
                item["excess_qty"] = (base_panels * 2) - prod_qty
            elif cat in ['서비스', '기타']:
                item["panels"] = 0
                item["is_bumper_applied"] = False
                item["base_panels"] = 0
                item["excess_qty"] = 0
            else:
                base_panels = math.ceil(prod_qty / bs) if bs > 0 and prod_qty > 0 else 0
                item["panels"] = base_panels + extra_panels
                item["is_bumper_applied"] = extra_panels > 0
                item["base_panels"] = base_panels
                item["excess_qty"] = (item["panels"] * bs) - prod_qty

        # Fetch custom layout from settings table (guaranteed persistence & handles separators)
        custom_layout = []
        try:
            settings_res = supabase.table("settings").select("value").eq("key", "product_sort_orders").execute()
            if settings_res.data:
                import json
                custom_layout = json.loads(settings_res.data[0]["value"])
        except Exception as e:
            print("Failed to load product_sort_orders from settings:", e)

        order_map = {}
        separators_to_add = []
        if custom_layout:
            for item in custom_layout:
                if isinstance(item, dict):
                    name = item.get("name")
                    s_order = item.get("sort_order", 999)
                    if name:
                        order_map[name] = s_order
                    if item.get("is_separator"):
                        separators_to_add.append({
                            "product_name": name or "--- 팀 구분선 ---",
                            "category": "기타",
                            "batch_size": 1,
                            "order_qty": 0,
                            "extra_qty": 0,
                            "required_qty": 0,
                            "carryover_qty": 0,
                            "production_qty": 0,
                            "panels": 0,
                            "base_panels": 0,
                            "excess_qty": 0,
                            "is_separator": True,
                            "sort_order": s_order,
                        })

        for item in parsed_items:
            name = item["product_name"]
            if name in order_map:
                item["sort_order"] = order_map[name]
            elif name in catalog_map:
                item["sort_order"] = catalog_map[name].get("sort_order", 999)
            else:
                item["sort_order"] = 999

        if separators_to_add:
            parsed_items.extend(separators_to_add)

        # Sort items strictly by sort_order ASC, then product_name ASC
        parsed_items.sort(key=lambda i: (i.get("sort_order", 999), i.get("product_name", "")))

        return {
            "status": "success",
            "items": parsed_items,
            "set_breakdowns": set_breakdowns
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse production excel: {str(e)}")


@app.post("/api/production/parse-shipment-excel")
async def api_parse_shipment_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        res = parse_shipment_notes_excel(contents)
        return res
    except Exception as e:
        print("Error parsing shipment excel:", e)
        raise HTTPException(status_code=500, detail=f"Failed to parse shipment excel: {str(e)}")


@app.get("/api/production/catalog")
def api_get_product_catalog():
    try:
        res = supabase.table("product_catalog").select("*").order("sort_order").order("name").execute()
        return {"status": "success", "data": res.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/production/catalog")
def api_save_product_catalog(payload: ProductCatalogPayload):
    try:
        data_to_save = {
            "name": payload.name,
            "category": payload.category,
            "batch_size": payload.batch_size,
            "min_bumper_qty": payload.min_bumper_qty,
            "is_confirmed": payload.is_confirmed,
            "parent_scone_name": payload.parent_scone_name,
            "oven_number": payload.oven_number if (payload.oven_number is not None and str(payload.oven_number).strip() != "") else "1",
            "heavy_cream_per_panel": payload.heavy_cream_per_panel if payload.heavy_cream_per_panel is not None else 0,
            "sort_order": payload.sort_order if payload.sort_order is not None else 999,
        }
        res = supabase.table("product_catalog").upsert(data_to_save, on_conflict="name").execute()
        return {"status": "success", "data": res.data}
    except Exception as e:
        print("Error saving product catalog item:", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/production/catalog/reorder")
def api_reorder_product_catalog(payload: ReorderCatalogPayload):
    try:
        import json
        orders_data = [item.dict() for item in payload.orders]
        # 1. Store in settings table for guaranteed persistence
        supabase.table("settings").upsert({"key": "product_sort_orders", "value": json.dumps(orders_data)}, on_conflict="key").execute()

        # 2. Also try updating product_catalog table column for product items
        for item in payload.orders:
            if not getattr(item, 'is_separator', False):
                try:
                    supabase.table("product_catalog").update({"sort_order": item.sort_order}).eq("name", item.name).execute()
                except Exception as ex:
                    print(f"Ignored catalog column update for {item.name}:", ex)

        return {"status": "success", "updated_count": len(payload.orders)}
    except Exception as e:
        print("Error reordering product catalog:", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/production/catalog/{name}")
def api_delete_product_catalog(name: str):
    try:
        res = supabase.table("product_catalog").delete().eq("name", name).execute()
        return {"status": "success", "data": res.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/production/sets")
def api_get_set_catalog():
    try:
        sets_res = supabase.table("set_catalog").select("*").order("name").execute()
        
        if not sets_res.data:
            return {"status": "success", "data": []}

        set_ids = [s["id"] for s in sets_res.data]
        items_res = supabase.table("set_items").select("*").in_("set_id", set_ids).execute()
        items_by_set = {}
        if items_res.data:
            for it in items_res.data:
                items_by_set.setdefault(it["set_id"], []).append({
                    "id": it.get("id"),
                    "set_id": it.get("set_id"),
                    "product_name": it.get("product_name"),
                    "quantity": it.get("quantity", 1)
                })

        result = []
        for s in sets_res.data:
            result.append({
                "id": s.get("id"),
                "set_name": s["name"],
                "description": s.get("description", ""),
                "is_confirmed": s.get("is_confirmed", True),
                "components": items_by_set.get(s["id"], [])
            })

        return {"status": "success", "data": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/production/sets")
def api_save_set_catalog(payload: SetCatalogPayload):
    try:
        target_name = payload.name or payload.set_name
        if not target_name:
            raise HTTPException(status_code=400, detail="Set name is required")

        set_res = supabase.table("set_catalog").upsert({
            "name": target_name,
            "description": payload.description or "",
            "is_confirmed": payload.is_confirmed,
        }, on_conflict="name").execute()

        if not set_res.data:
            raise Exception("Failed to upsert set_catalog")

        set_id = set_res.data[0]["id"]

        supabase.table("set_items").delete().eq("set_id", set_id).execute()

        comp_list = payload.items if payload.items is not None else (payload.components or [])
        if comp_list:
            items_to_insert = [
                {
                    "set_id": set_id,
                    "product_name": comp.product_name,
                    "quantity": comp.quantity
                }
                for comp in comp_list
            ]
            supabase.table("set_items").insert(items_to_insert).execute()

        return {"status": "success", "data": set_res.data[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/production/sets/{name}")
def api_delete_set_catalog(name: str):
    try:
        from urllib.parse import unquote
        clean_name = unquote(name).strip()
        
        # Search by decoded name, original name, or id
        set_res = supabase.table("set_catalog").select("id").eq("name", clean_name).execute()
        if not set_res.data:
            set_res = supabase.table("set_catalog").select("id").eq("name", name).execute()
        if not set_res.data:
            try:
                set_res = supabase.table("set_catalog").select("id").eq("id", clean_name).execute()
            except Exception:
                pass

        if set_res.data:
            for s in set_res.data:
                s_id = s["id"]
                supabase.table("set_items").delete().eq("set_id", s_id).execute()
                supabase.table("set_catalog").delete().eq("id", s_id).execute()
        return {"status": "success"}
    except Exception as e:
        print("Set delete error:", e)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/production/save-summary")
def api_save_production_summary(payload: Dict[str, Any]):
    try:
        record_date = payload.get("record_date")
        items = payload.get("items", [])

        if not record_date:
            raise HTTPException(status_code=400, detail="record_date is required")

        supabase.table("production_records").delete().eq("record_date", record_date).execute()

        rows_to_insert = []
        for item in items:
            rows_to_insert.append({
                "record_date": record_date,
                "product_name": item["product_name"],
                "category": item["category"],
                "batch_size": item.get("batch_size", 8),
                "order_qty": item.get("order_qty", 0),
                "extra_qty": item.get("extra_qty", 0),
                "required_qty": item.get("required_qty", 0),
                "carryover_qty": item.get("carryover_qty", 0),
                "production_qty": item.get("production_qty", 0),
                "panels": item.get("panels", 0),
                "is_bumper_applied": item.get("is_bumper_applied", False),
                "excess_qty": item.get("excess_qty", 0),
                "min_bumper_qty": item.get("min_bumper_qty", 2),
                "is_confirmed": item.get("is_confirmed", True),
                "parent_scone_name": item.get("parent_scone_name"),
            })

        if rows_to_insert:
            supabase.table("production_records").insert(rows_to_insert).execute()

        return {"status": "success", "inserted_count": len(rows_to_insert)}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
