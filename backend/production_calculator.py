import openpyxl
import math
import re
import io
from typing import List, Dict, Any

def clean_product_name(name: str) -> str:
    """Strips invoice-ordering prefixes like '---', '-', spaces from product names."""
    if not name:
        return ""
    cleaned = re.sub(r'^[-_\s]+', '', str(name)).strip()
    return cleaned

def normalize_scone_key(name: str) -> str:
    if not name:
        return ""
    s = re.sub(r'\[.*?\]', '', str(name))
    s = re.sub(r'\(.*?\)', '', s)
    for kw in ['스틱', '스콘', '3팩', '1팩', '하프팩', '미니큐브', '미니쉐이크', '서비스', '팩']:
        s = s.replace(kw, '')
    return re.sub(r'\s+', '', s).strip().lower()

def parse_production_excel(file_contents: bytes, set_catalog_map: Dict[str, Any] = None, catalog_dict: Dict[str, Any] = None) -> Dict[str, Any]:
    """
    Parses EasyAdmin production order excel file bytes.
    Detects Product Name and Option columns.
    Combines product name and option text into raw full item names.
    Decomposes registered Set Products into single scone components.
    Aggregates order quantities by product name and categorizes items.
    """
    workbook = openpyxl.load_workbook(filename=io.BytesIO(file_contents), data_only=True)
    target_sheet = workbook.active

    # 1. Locate header row and target columns dynamically
    header_row_idx = None
    product_col_idx = None
    option_col_idx = None
    qty_col_idx = None

    for r in range(1, min(10, target_sheet.max_row + 1)):
        row_str = [str(target_sheet.cell(r, c).value or '').strip() for c in range(1, min(25, target_sheet.max_column + 1))]
        for idx, val in enumerate(row_str):
            if val in ['상품명', '품목명']:
                product_col_idx = idx + 1
            elif val == '공급처상품명' and not product_col_idx:
                product_col_idx = idx + 1
            elif val in ['옵션', '옵션명', '품목옵션', '공급처옵션명', '옵션정보']:
                option_col_idx = idx + 1
            elif val in ['수량', '주문수량', '품목수량']:
                qty_col_idx = idx + 1

        if product_col_idx and qty_col_idx:
            header_row_idx = r
            break

    # Fallback default positions if headers not found by exact name
    if not product_col_idx:
        product_col_idx = 5  # Column E in summary format
    if not option_col_idx:
        option_col_idx = 6   # Column F in summary format
    if not qty_col_idx:
        qty_col_idx = 17     # Column Q in summary format
    if not header_row_idx:
        header_row_idx = 1

    # 2. Aggregate order quantities by clean product name & handle Set / Halfpack / MiniShake / Stick / Service / Misc
    product_orders: Dict[str, Dict[str, Any]] = {}
    set_breakdowns: List[Dict[str, Any]] = []

    set_map = set_catalog_map or {}

    def get_category_info(clean_name: str):
        c_lower = clean_name.lower()
        is_service = '서비스' in clean_name
        is_misc = any(k in c_lower for k in ['요프', '요거트', 'opp', '대파', '피넛', '스무스', '크런치', '스타터팩', '이매진', '유크림', '기타'])
        is_stick = not is_service and not is_misc and '스틱' in clean_name
        is_shake = not is_service and not is_misc and not is_stick and ('미니쉐이크' in clean_name or '쉐이크' in clean_name)
        is_halfpack = not is_service and not is_misc and not is_stick and not is_shake and ('하프팩' in clean_name or '미니큐브' in clean_name)
        is_bar = not is_service and not is_misc and '바' in clean_name and ('스콘' not in clean_name or '바스콘' in clean_name or clean_name.endswith('바'))
        
        if is_service:
            return '서비스', 1
        elif is_misc:
            return '기타', 1
        elif is_stick:
            return '스틱', 9  # 9 packs per panel (1 EasyAdmin qty = 3 packs)
        elif is_shake:
            return '미니쉐이크', 4  # 4 bags per panel (18 pieces per bag)
        elif is_halfpack:
            return '미니큐브', 2  # 2 bags per panel (36 pieces per bag)
        elif is_bar:
            return '바', 10
        else:
            return '삼각', 8

    for r in range(header_row_idx + 1, target_sheet.max_row + 1):
        raw_name = target_sheet.cell(r, product_col_idx).value
        raw_option = target_sheet.cell(r, option_col_idx).value if option_col_idx else None
        raw_qty = target_sheet.cell(r, qty_col_idx).value

        if not raw_name:
            continue

        clean_name = clean_product_name(str(raw_name))
        if not clean_name:
            continue

        # Option text handling
        opt_str = str(raw_option).strip() if raw_option and str(raw_option).strip() not in ['None', 'none', '없음', '기본', '단품'] else ''
        full_clean_name = f"{clean_name} ({opt_str})" if opt_str else clean_name

        try:
            qty = int(raw_qty) if raw_qty is not None else 0
        except (ValueError, TypeError):
            qty = 0

        if qty <= 0:
            continue

        # Check if item matches a registered Set Product (exact match only for manager control & safety)
        if full_clean_name in set_map or clean_name in set_map:
            s_name = full_clean_name if full_clean_name in set_map else clean_name
            set_info = set_map[s_name]
            if isinstance(set_info, list):
                components = set_info
            elif isinstance(set_info, dict):
                components = set_info.get("items", []) or set_info.get("components", [])
            else:
                components = []
            
            set_breakdowns.append({
                "set_name": s_name,
                "set_order_qty": qty,
                "components": components,
            })

            # Multiply set order qty by component qty and add to component scones
            for comp in components:
                comp_name = comp["product_name"]
                comp_qty = comp.get("quantity", 1) * qty
                category, batch_size = get_category_info(comp_name)

                if comp_name not in product_orders:
                    product_orders[comp_name] = {
                        "raw_name": comp_name,
                        "order_qty": 0,
                        "category": category,
                        "batch_size": batch_size,
                        "is_set_component": True
                    }
                product_orders[comp_name]["order_qty"] += comp_qty
        else:
            # Single Scone / Halfpack / MiniShake / Stick / Service / Misc Item
            category, batch_size = get_category_info(full_clean_name)
            
            # If a single stick item name contains '3팩', each order unit contains 3 packs of sticks
            multiplier = 3 if (category == '스틱' and '3팩' in full_clean_name) else 1
            calculated_qty = qty * multiplier

            if full_clean_name not in product_orders:
                product_orders[full_clean_name] = {
                    "raw_name": f"{raw_name} [{opt_str}]" if opt_str else str(raw_name).strip(),
                    "order_qty": 0,
                    "category": category,
                    "batch_size": batch_size,
                    "is_set_component": True if category == '스틱' else False
                }
            product_orders[full_clean_name]["order_qty"] += calculated_qty

    # 3. Calculate batches, panels, bumper, and excess cleanly
    result = []
    for clean_name, item in product_orders.items():
        category = item["category"]
        batch_size = item["batch_size"]
        order_qty = item["order_qty"]
        is_set_comp = item.get("is_set_component", False)

        min_bumper_qty = 2
        extra_qty = 0
        carryover_qty = 0

        required_qty = order_qty + extra_qty
        production_qty = max(0, required_qty - carryover_qty)

        if category == '스틱':
            base_panels = math.ceil(production_qty / 9.0) if production_qty > 0 else 0
            excess_qty = (base_panels * 9) - production_qty
            is_bumper_applied = False
            panels = base_panels
        elif category == '미니쉐이크':
            base_panels = math.ceil(production_qty / 4.0) if production_qty > 0 else 0
            excess_qty = (base_panels * 4) - production_qty
            is_bumper_applied = False
            panels = base_panels
        elif category == '미니큐브':
            base_panels = math.ceil(production_qty / 2.0) if production_qty > 0 else 0
            excess_qty = (base_panels * 2) - production_qty
            is_bumper_applied = False
            panels = base_panels
        elif category in ['서비스', '기타']:
            base_panels = 0
            panels = 0
            is_bumper_applied = False
            excess_qty = 0
        else:
            base_panels = math.ceil(production_qty / float(batch_size)) if production_qty > 0 and batch_size > 0 else 0
            base_excess = (base_panels * batch_size) - production_qty

            if production_qty > 0 and base_excess < min_bumper_qty:
                panels = base_panels + 1
                is_bumper_applied = True
                excess_qty = (panels * batch_size) - production_qty
            else:
                panels = base_panels
                is_bumper_applied = False
                excess_qty = base_excess

        result.append({
            "product_name": clean_name,
            "raw_name": item["raw_name"],
            "category": category,
            "batch_size": batch_size,
            "min_bumper_qty": min_bumper_qty,
            "parent_scone_name": None,
            "order_qty": order_qty,
            "extra_qty": extra_qty,
            "required_qty": required_qty,
            "carryover_qty": carryover_qty,
            "production_qty": production_qty,
            "base_panels": base_panels,
            "panels": panels,
            "is_bumper_applied": is_bumper_applied,
            "excess_qty": excess_qty,
            "halfpack_order_qty": order_qty if category == '미니큐브' else 0,
            "halfpack_panels": round(order_qty / 2.0, 1) if category == '미니큐브' else 0,
            "is_set_component": is_set_comp,
        })

    return {
        "status": "success",
        "items": result,
        "set_breakdowns": set_breakdowns
    }


def parse_shipment_notes_excel(file_contents: bytes) -> Dict[str, Any]:
    """
    Parses an order shipment excel file to calculate:
    - Per-order quantities of Greek Yogurt ('-----GREEK YOGURT') and YOF ('------YOF6팩')
    - Frequency distributions for yogurt and YOF
    - Combined shipments (합배송: identical tracking number with different order IDs), extracting 주문자명 and 주문자전화번호
    - Jeju customer candidates
    - Total order counts
    """
    workbook = openpyxl.load_workbook(filename=io.BytesIO(file_contents), data_only=True)
    target_sheet = workbook.active

    # Locate headers
    header_row_idx = 1
    order_id_col = None
    tracking_col = None
    buyer_name_col = None
    buyer_phone_col = None
    receiver_name_col = None
    receiver_phone_col = None
    product_col = None
    option_col = None
    qty_col = None
    address_col = None

    max_cols = max(target_sheet.max_column, 50)

    for r in range(1, min(10, target_sheet.max_row + 1)):
        row_cells = [str(target_sheet.cell(r, c).value or '').strip() for c in range(1, max_cols + 1)]
        
        found_product = False
        found_qty = False

        for idx, val in enumerate(row_cells):
            col_num = idx + 1
            if not val:
                continue
            clean_val = val.replace(' ', '').replace('\n', '').replace('\r', '').lower()

            # 1. Tracking column
            if any(k in clean_val for k in ['송장번호', '운송장번호', '송장no', '운송장no', '택배송장']) or clean_val in ['송장', '운송장']:
                if not tracking_col: tracking_col = col_num

            # 2. Order ID column
            elif any(k in clean_val for k in ['주문번호', '발주번호', '주문고유번호', '배송번호', '쇼핑몰주문번호', '원주문번호']):
                if not order_id_col: order_id_col = col_num

            # 3. Buyer Phone column (주문자 전화번호 / 연락처)
            elif any(b in clean_val for b in ['주문자', '구매자', '주문고객', '보내는', '송하인']) and any(p in clean_val for p in ['전화', '연락처', '휴대폰', '핸드폰', '이동전화', 'tel', 'phone', 'hp']):
                if not buyer_phone_col: buyer_phone_col = col_num

            # 4. Buyer Name column (주문자명 / 구매자명)
            elif any(b in clean_val for b in ['주문자명', '구매자명', '주문자이름', '구매자이름', '주문고객명', '주문자인', '보내는사람', '보내는분', '송하인명', '송하인']) or clean_val in ['주문자', '구매자', '주문고객']:
                if not any(ex in clean_val for ex in ['전화', '연락처', '휴대폰', '핸드폰', '주소', '우편', '이메일', 'id', '아이디', '수취', '수령', '받는']):
                    if not buyer_name_col: buyer_name_col = col_num

            # 5. Receiver Phone column (수취인 전화번호 / 연락처)
            elif any(r_kw in clean_val for r in ['수취인', '수령인', '받는사람', '받는분', '수하인'] for r_kw in [r]) and any(p in clean_val for p in ['전화', '연락처', '휴대폰', '핸드폰', '이동전화', 'tel', 'phone', 'hp']):
                if not receiver_phone_col: receiver_phone_col = col_num

            # 6. Receiver Name column (수취인명 / 수령인)
            elif any(r_kw in clean_val for r in ['수취인명', '수령인명', '받는분', '받는사람', '수하인명'] for r_kw in [r]) or clean_val in ['수취인', '수령인', '수하인']:
                if not any(ex in clean_val for ex in ['전화', '연락처', '휴대폰', '핸드폰', '주소', '우편']):
                    if not receiver_name_col: receiver_name_col = col_num

            # 7. Product column
            elif any(p in clean_val for p in ['상품명', '품목명', '공급처상품명']):
                if not product_col: product_col = col_num
                found_product = True

            # 8. Option column
            elif any(o in clean_val for o in ['옵션', '옵션명', '품목옵션', '옵션정보', '공급처옵션명']):
                if not option_col: option_col = col_num

            # 9. Qty column
            elif any(q in clean_val for q in ['수량', '주문수량', '품목수량']):
                if not qty_col: qty_col = col_num
                found_qty = True

            # 10. Address column
            elif any(a in clean_val for a in ['주소', '수취인주소', '배송지', '배송지주소']):
                if not address_col: address_col = col_num

        if found_product and found_qty:
            header_row_idx = r
            break

    if not product_col: product_col = 5
    if not option_col: option_col = 6
    if not qty_col: qty_col = 17

    print(f"[ShipmentParser] Headers: header_row={header_row_idx}, tracking={tracking_col}, order_id={order_id_col}, buyer_name={buyer_name_col}, buyer_phone={buyer_phone_col}, receiver_name={receiver_name_col}, receiver_phone={receiver_phone_col}, prod={product_col}, qty={qty_col}")

    order_yogurt_map: Dict[str, int] = {}
    order_yof_map: Dict[str, int] = {}
    all_order_keys = set()
    jeju_candidates = []
    
    # Map tracking_number -> { "orders": set(), "buyer_name": str, "buyer_phone": str }
    tracking_groups: Dict[str, Dict[str, Any]] = {}

    for r in range(header_row_idx + 1, target_sheet.max_row + 1):
        raw_order_id = str(target_sheet.cell(r, order_id_col).value or '').strip() if order_id_col else None
        raw_tracking = str(target_sheet.cell(r, tracking_col).value or '').strip() if tracking_col else None
        raw_buyer_name = str(target_sheet.cell(r, buyer_name_col).value or '').strip() if buyer_name_col else ''
        raw_buyer_phone = str(target_sheet.cell(r, buyer_phone_col).value or '').strip() if buyer_phone_col else ''
        raw_receiver = str(target_sheet.cell(r, receiver_name_col).value or '').strip() if receiver_name_col else None
        raw_receiver_phone = str(target_sheet.cell(r, receiver_phone_col).value or '').strip() if receiver_phone_col else ''
        raw_product = str(target_sheet.cell(r, product_col).value or '').strip()
        raw_option = str(target_sheet.cell(r, option_col).value or '').strip() if option_col else ''
        raw_qty_val = target_sheet.cell(r, qty_col).value
        raw_addr = str(target_sheet.cell(r, address_col).value or '').strip() if address_col else ''

        if not raw_product and not raw_option and not raw_tracking:
            continue

        try:
            qty = int(float(str(raw_qty_val).replace(',', '').strip()))
        except:
            qty = 1

        order_key = raw_order_id or raw_receiver or f"order_{r}"
        all_order_keys.add(order_key)

        # Track combined shipments by tracking number
        if raw_tracking and raw_tracking not in ['-', '', '0', 'None', 'nan']:
            if raw_tracking not in tracking_groups:
                tracking_groups[raw_tracking] = {
                    "orders": set(),
                    "buyer_name": '',
                    "buyer_phone": '',
                }

            # Prioritize buyer_name, fallback to receiver name if buyer name missing
            best_name = raw_buyer_name or raw_receiver or ''
            best_phone = raw_buyer_phone or raw_receiver_phone or ''

            if best_name and (not tracking_groups[raw_tracking]["buyer_name"] or tracking_groups[raw_tracking]["buyer_name"] in ['이름 없음', '-']):
                tracking_groups[raw_tracking]["buyer_name"] = best_name
            if best_phone and (not tracking_groups[raw_tracking]["buyer_phone"] or tracking_groups[raw_tracking]["buyer_phone"] in ['-', '']):
                tracking_groups[raw_tracking]["buyer_phone"] = best_phone
            
            if raw_order_id:
                tracking_groups[raw_tracking]["orders"].add(raw_order_id)
            else:
                tracking_groups[raw_tracking]["orders"].add(f"row_{r}")


        full_prod_text = f"{raw_product} {raw_option}".upper()

        # Check Greek Yogurt ('-----GREEK YOGURT' or 'GREEK YOGURT' or '그릭요거트')
        is_yogurt = ('GREEK YOGURT' in full_prod_text or 'GREEKYOGURT' in full_prod_text or '그릭요거트' in full_prod_text or '그릭 요거트' in full_prod_text)
        if is_yogurt:
            order_yogurt_map[order_key] = order_yogurt_map.get(order_key, 0) + qty

        # Check YOF ('------YOF6팩' or 'YOF6' or 'YOF-6' or '요프')
        is_yof = ('YOF6' in full_prod_text or 'YOF-6' in full_prod_text or 'YOF 6' in full_prod_text or '요프6' in full_prod_text or '요프 6' in full_prod_text)
        if is_yof:
            order_yof_map[order_key] = order_yof_map.get(order_key, 0) + qty

        # Check Jeju address
        if raw_addr and ('제주' in raw_addr or '서귀포' in raw_addr):
            if raw_receiver and raw_receiver not in jeju_candidates:
                jeju_candidates.append(raw_receiver)

    # Frequency distributions
    greek_distribution = {i: 0 for i in range(1, 11)}
    for order_key, total_qty in order_yogurt_map.items():
        if 1 <= total_qty <= 10:
            greek_distribution[total_qty] += 1
        elif total_qty > 10:
            greek_distribution[10] += 1

    yof_distribution = {1: 0, 2: 0, 3: 0}
    for order_key, total_qty in order_yof_map.items():
        if 1 <= total_qty <= 3:
            yof_distribution[total_qty] += 1
        elif total_qty > 3:
            yof_distribution[3] += 1

    # Extract combined shipments (different order IDs with same tracking number)
    combined_shipments = []
    for tracking_num, info in tracking_groups.items():
        if len(info["orders"]) > 1:
            combined_shipments.append({
                "tracking_number": tracking_num,
                "buyer_name": info["buyer_name"] or '이름 없음',
                "buyer_phone": info["buyer_phone"] or '-',
                "order_count": len(info["orders"]),
                "order_ids": list(info["orders"]),
            })

    # Sort combined shipments by buyer name
    combined_shipments.sort(key=lambda x: x["buyer_name"])

    return {
        "status": "success",
        "greek_yogurt": greek_distribution,
        "yof": yof_distribution,
        "combined_shipments": combined_shipments,
        "total_orders": len(all_order_keys),
        "jeju_candidates": jeju_candidates,
    }
