import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import HTTPException
from fastapi.responses import FileResponse

def export_report_excel(report: dict, rows: list, template_path: str, upload_dir: str):
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Template Excel not found")
        
    export_filename = f"vat_report_{report['year']}_{report['quarter']}Q.xlsx"
    export_path = os.path.join(upload_dir, export_filename)
    shutil.copyfile(template_path, export_path)
    
    wb = openpyxl.load_workbook(export_path, data_only=False)
    sheet = wb.active
    
    sheet.title = f"{report['year']}년 {report['quarter']}분기"
    sheet.cell(row=1, column=2).value = f"{report['year']}년 {report['quarter']}분기 판매금액 결제수단 별 정리"
    sheet.cell(row=2, column=5).value = "메모"
    
    for r_data in rows:
        excel_row = r_data.get("excel_row")
        if not excel_row:
            continue
            
        classification = r_data.get("classification")
        brand = r_data.get("brand")
        amount = r_data.get("amount")
        memo = r_data.get("memo")
        
        if classification == "무통장입금" and brand in ["머드스콘", "오터", "위시", "위시\n(그릭요거트)", "위시 (그릭요거트)"]:
            sheet.cell(row=excel_row, column=6).value = r_data.get("bank_income")
            sheet.cell(row=excel_row, column=7).value = r_data.get("bank_expense")
            sheet.cell(row=excel_row, column=8).value = r_data.get("bank_toss_receipt")
            if brand == "오터":
                sheet.cell(row=excel_row, column=9).value = r_data.get("bank_koces_receipt")
            if amount is not None:
                sheet.cell(row=excel_row, column=4).value = amount
        elif brand == "홈택스 현금영수증":
            sheet.cell(row=excel_row, column=3).value = f"{report['year']}년 {report['quarter']}분기"
            if amount is not None:
                sheet.cell(row=excel_row, column=4).value = amount
        else:
            is_formula_row = False
            amount_val = sheet.cell(row=excel_row, column=4).value
            if isinstance(amount_val, str) and amount_val.startswith('='):
                is_formula_row = True
            elif classification == '합계' or brand == '판매금액 총합':
                is_formula_row = True
                
            if not is_formula_row and amount is not None:
                sheet.cell(row=excel_row, column=4).value = amount
                
        sheet.cell(row=excel_row, column=5).value = memo if memo else ""
                 
    sheet.print_area = None
    sheet.row_breaks = []
    sheet.col_breaks = []
    
    if hasattr(sheet, 'views') and sheet.views and sheet.views.sheetView:
        sheet.views.sheetView[0].view = 'normal'
        sheet.views.sheetView[0].showGridLines = True
    else:
        sheet.sheet_view.view = 'normal'
        
    font_family = "Malgun Gothic"
    
    title_font = Font(name=font_family, size=15, bold=True, color="0F172A")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    normal_font = Font(name=font_family, size=10, color="334155")
    sum_row_font = Font(name=font_family, size=10, bold=True, color="1E293B")
    total_row_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    total_amount_font = Font(name=font_family, size=11, bold=True, color="10B981")
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    sum_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    total_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    hometax_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(style='thin', color='CBD5E1')
    sum_top_side = Side(style='thin', color='94A3B8')
    sum_bottom_side = Side(style='medium', color='64748B')
    total_bottom_side = Side(style='double', color='0F172A')
    
    normal_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    sum_border = Border(left=thin_side, right=thin_side, top=sum_top_side, bottom=sum_bottom_side)
    total_border = Border(left=thin_side, right=thin_side, top=sum_top_side, bottom=total_bottom_side)
    
    sheet.row_dimensions[1].height = 35
    sheet.row_dimensions[2].height = 26
    
    title_cell = sheet.cell(row=1, column=2)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    for col_idx in range(1, 6):
        cell = sheet.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = normal_border
        
    for r in range(3, 81):
        sheet.row_dimensions[r].height = 20
        classification = str(sheet.cell(row=r, column=3).value or "").strip()
        
        is_sum = (classification == "합계")
        is_total = (r == 79)
        is_hometax = (r == 80)
        
        for col_idx in range(1, 6):
            cell = sheet.cell(row=r, column=col_idx)
            
            if col_idx in [1, 2, 3]:
                cell.alignment = center_align
            elif col_idx == 4:
                cell.alignment = right_align
            else:
                cell.alignment = left_align
                
            if is_total:
                cell.fill = total_fill
                if col_idx == 1:
                    cell.value = "총 매출 합계 (판매금액 총합)"
                    cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
                elif col_idx == 4:
                    cell.value = "=SUM(D3,D8,D14,D20,D26,D33,D37,D38,D42,D48,D54,D58,D62,D63,D64,D69,D75,D76,D77,D78)"
                    cell.font = total_amount_font
                else:
                    cell.font = total_row_font
                cell.border = total_border
            elif is_sum:
                cell.fill = sum_fill
                cell.font = sum_row_font
                cell.border = sum_border
            elif is_hometax:
                cell.fill = hometax_fill
                cell.font = normal_font
                cell.border = normal_border
            else:
                cell.font = normal_font
                cell.border = normal_border
                
            if col_idx == 4 and not is_total:
                cell.number_format = '#,##0'
                
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 24
    sheet.column_dimensions['C'].width = 26
    sheet.column_dimensions['D'].width = 22
    sheet.column_dimensions['E'].width = 30
    
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['G'].width = 18
    sheet.column_dimensions['H'].width = 18
    sheet.column_dimensions['I'].width = 18
    
    wb.save(export_path)
    return FileResponse(
        path=export_path, 
        filename=export_filename, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
